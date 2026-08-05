#!/usr/bin/env python3
"""
sync_data.py — EFIKA Portal Consalud
Descarga "Control Consolidado Servicios.xlsx" de SharePoint
via Microsoft Graph API y genera data.js con todos los datos del portal.

Uso:
  python sync_data.py                      # descarga SharePoint → genera data.js
  python sync_data.py --push               # + git push a GitHub Pages
  python sync_data.py --local archivo.xlsx # usar Excel local (sin SharePoint)

Variables de entorno requeridas (o en archivo .env):
  AZURE_TENANT_ID      = bdabe570-1ae7-4cd4-8deb-7243d62c4520
  AZURE_CLIENT_ID      = 1a58b6b6-fd06-4ac6-9469-54f1abd536ca   # App "Agente EFIKA"
  AZURE_CLIENT_SECRET  = <secreto creado en Azure Portal>
"""

import sys, os, json, subprocess, tempfile
from datetime import datetime
from pathlib import Path

# ── CARGAR .env si existe ─────────────────────────────────────────────────────
env_file = Path(__file__).parent / ".env"
if env_file.exists():
    for line in env_file.read_text(encoding="utf-8").splitlines():
        if "=" in line and not line.startswith("#"):
            k, v = line.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip())

# ── CONFIG ────────────────────────────────────────────────────────────────────
TENANT_ID     = os.environ.get("AZURE_TENANT_ID",     "bdabe570-1ae7-4cd4-8deb-7243d62c4520")
CLIENT_ID     = os.environ.get("AZURE_CLIENT_ID",     "1a58b6b6-fd06-4ac6-9469-54f1abd536ca")
CLIENT_SECRET = os.environ.get("AZURE_CLIENT_SECRET", "")

# Ruta del archivo en SharePoint
# SP_LIBRARY   = nombre de la biblioteca en Graph API (primer segmento del path)
# SP_FILE_PATH = ruta del archivo DENTRO de la biblioteca (sin el prefijo de biblioteca)
SP_SITE      = "efika2022.sharepoint.com:/sites/EFIKA:"
SP_LIBRARY   = "Documentos"
SP_FILE_PATH = "CLIENTES ACTUALES/CONSALUD/Control De Servicio/Control Consolidado Servicios/Control Consolidado Servicios.xlsx"


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS DE CONVERSIÓN
# ══════════════════════════════════════════════════════════════════════════════

def to_num(val):
    """Celda Excel → float (0 si no aplica)."""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace("$", "").replace(" ", "").replace(".", "").replace(",", ".").strip()
    try:
        return float(s)
    except ValueError:
        return 0

def to_int(val):
    return int(round(to_num(val)))

def to_ddmmyyyy(val):
    """Fecha → 'DD/MM/YYYY'."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y")
    s = str(val).strip()
    if len(s) == 10:
        if s[2] == "/":
            return s
        if s[4] == "-":
            return f"{s[8:10]}/{s[5:7]}/{s[0:4]}"
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s[:10], fmt[:10]).strftime("%d/%m/%Y")
        except (ValueError, IndexError):
            continue
    return s

def to_yyyymm(val):
    """Fecha → 'YYYY-MM'."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m")
    s = str(val).strip()
    if len(s) >= 7 and s[4] == "-":
        return s[:7]
    if len(s) == 10 and s[2] == "/":
        return f"{s[6:10]}-{s[3:5]}"
    return s

_MESES_ES = {"ene":1,"feb":2,"mar":3,"abr":4,"may":5,"jun":6,
             "jul":7,"ago":8,"sep":9,"set":9,"oct":10,"nov":11,"dic":12}

def _mes_texto(s):
    """'junio 2022' / 'jun-22' / 'Junio-2022' -> '2022-06-01'."""
    import re as _re
    t = str(s).lower().strip()
    mm = None
    for k, v in _MESES_ES.items():
        if k in t:
            mm = v; break
    if not mm:
        return ""
    y = _re.search(r"(20\d{2})", t)
    if y:
        yy = int(y.group(1))
    else:
        y2 = _re.search(r"(\d{2})\s*$", t)
        if not y2:
            return ""
        yy = 2000 + int(y2.group(1))
    return f"{yy}-{mm:02d}-01"

def to_yyyymmdd(val):
    """Fecha → 'YYYY-MM-DD' (para resumen.m, compatible con .slice(0,7))."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if len(s) >= 10 and s[4] == "-":
        return s[:10]
    if len(s) == 10 and s[2] == "/":
        return f"{s[6:10]}-{s[3:5]}-{s[0:2]}"
    return s

def clean(val):
    return "" if val is None else str(val).strip()


# ── Matching flexible de nombres (acentos/mayúsculas/tokens) ──────────────────
import unicodedata

def norm_txt(s):
    s = str(s or "").strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return " ".join(s.split())

def _tokens(s):
    return set(norm_txt(s).split())

class NormRow(dict):
    """dict con .get() tolerante: exacto -> normalizado -> tokens contenidos."""
    def get(self, key, default=None):
        if key in self:
            v = dict.get(self, key)
            if v is not None and str(v).strip() != "":
                return v
        nk = norm_txt(key)
        for k in self.keys():
            if norm_txt(k) == nk:
                v = dict.get(self, k)
                if v is not None and str(v).strip() != "":
                    return v
        tk = _tokens(key)
        if tk:
            for k in self.keys():
                if tk <= _tokens(k):
                    v = dict.get(self, k)
                    if v is not None and str(v).strip() != "":
                        return v
        return default



def sheet_as_dicts(wb, *possible_names):
    """
    Abre la primera hoja cuyo nombre coincida (sin distinguir acentos/mayúsculas),
    detecta la fila de encabezado y retorna lista de dicts.
    """
    ws = None
    # 1) match exacto (normalizado: sin acentos ni mayusculas)
    for target in possible_names:
        tn = norm_txt(target)
        for name in wb.sheetnames:
            if norm_txt(name) == tn:
                ws = wb[name]
                break
        if ws:
            break
    # 2) match por tokens contenidos (ej: 'TELEFONIA FIJA' vs 'Servicios Telefonia Fija')
    if ws is None:
        for target in possible_names:
            tt = _tokens(target)
            for name in wb.sheetnames:
                nt = _tokens(name)
                if nt and (nt <= tt or tt <= nt):
                    ws = wb[name]
                    break
            if ws:
                break

    if ws is None:
        names_str = ", ".join(wb.sheetnames)
        print(f"    ⚠  Ninguna de las hojas {list(possible_names)} encontrada. Disponibles: {names_str}")
        return []

    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return []

    # Fila de encabezado = primera fila con ≥3 celdas no vacías
    hdr_idx = 0
    for i, row in enumerate(all_rows):
        if sum(1 for c in row if c is not None and str(c).strip()) >= 3:
            hdr_idx = i
            break

    headers = [clean(h) for h in all_rows[hdr_idx]]

    result = []
    for row in all_rows[hdr_idx + 1:]:
        if not any(c is not None for c in row):
            continue
        result.append(NormRow({headers[j]: row[j] for j in range(min(len(headers), len(row)))}))

    return result


# ══════════════════════════════════════════════════════════════════════════════
# PROCESADORES POR HOJA
# ══════════════════════════════════════════════════════════════════════════════

def proc_resumen(wb):
    """Hoja 'Resumen Ahorros' → D.resumen[]
    Schema: {m, total, mv, sms, pv, fi, ms, ah}
    - m:     'YYYY-MM-DD' (slice(0,7) → 'YYYY-MM' para labels de gráficos)
    - total: ahorro total del mes (= GESTIÓN EFIKA o suma categorías)
    - mv/sms/pv/fi/ms: ahorro por categoría
    - ah:    alias de total (usado por chatbot)
    """
    rows = sheet_as_dicts(wb, "Resumen Ahorros", "Resumen", "Ahorros Totales Servicios")
    out = []
    for r in rows:
        mes_val = r.get("FECHA AHORRO") or r.get("Mes") or r.get("mes") or (list(r.values())[0] if r else None)
        if not isinstance(mes_val, datetime) and not mes_val:
            continue
        m = to_yyyymmdd(mes_val)
        if not m or m.startswith("None"):
            continue
        if len(m) < 7 or m[4] != "-":       # texto tipo "junio 2022" / "jun-22"
            m = _mes_texto(m)
            if not m:
                continue

        mv  = to_int(r.get("Ahorro S.Movil")   or r.get("Ahorro S. Movil")  or r.get("MÓVILES") or 0)
        sms = to_int(r.get("Ahorro S. SMS")    or r.get("Ahorro S.SMS")     or 0)
        pv  = to_int(r.get("Ahorro S.Privado") or r.get("Ahorro S. Privado")or r.get("Servicios Privados en $") or 0)
        fi  = to_int(r.get("Ahorro T Fija")    or r.get("Ahorro T. Fija")   or r.get("TELEFONÍA IP") or 0)
        ms  = to_int(r.get("Ahorro Microsoft") or 0)
        tp  = to_int(r.get("Total Pack") or r.get("Ahorro Totalpack") or r.get("Ahorro TotalPack") or 0)
        ky  = to_int(r.get("Ahorro Kyochera")  or r.get("Ahorro Kyocera")   or 0)

        # GESTIÓN EFIKA = total mensual de ahorros gestionados por EFIKA
        sa  = to_int(r.get("Ahorro Seguridad America") or r.get("Ahorro Seguridad Medical") or 0)

        # En Consalud la columna "Ahorro Acumulado GESTION EFIKA" es ACUMULADA:
        # no sirve como total mensual. Se usa la suma de categorias del mes.
        total = mv + sms + pv + fi + ms + tp + ky + sa
        if total == 0:
            continue

        out.append({"m": m, "total": total, "mv": mv, "sms": sms, "pv": pv,
                    "fi": fi, "ms": ms, "tp": tp, "sa": sa, "ah": total})

    out.sort(key=lambda x: x["m"])
    print(f"    ✓ Resumen Ahorros: {len(out)} meses")
    return out


def proc_privados(wb):
    """Hoja 'Servicios Privados' → D.privados[]
    Schema: {f, net, fo, pr, desc, uf, st}
    """
    rows = sheet_as_dicts(wb, "Servicios Privados")
    out = []
    for r in rows:
        f   = to_ddmmyyyy(r.get("Fecha emisión") or r.get("Fecha Emisión") or r.get("Fecha") or "")
        net = to_int(r.get("MONTO NETO") or r.get("Monto Neto") or r.get("$ NETO") or r.get("NETO") or 0)
        if net == 0:
            continue
        out.append({
            "f":    f,
            "net":  net,
            "fo":   clean(r.get("Factura Folio") or r.get("Folio") or r.get("Numero Factura") or r.get("Nro Factura") or ""),
            "pr":   clean(r.get("Proveedor") or ""),
            "desc": clean(r.get("Descripción") or r.get("Descripcion") or r.get("Detalle Facturación") or ""),
            "uf":   round(to_num(r.get("Valor UF NETO") or r.get("Valor UF") or r.get("MONTO UF") or 0), 4),
            "st":   clean(r.get("Status") or r.get("status") or r.get("ESTADO") or r.get("APROBADA / RECHAZADA") or ""),
        })
    print(f"    ✓ Servicios Privados: {len(out)} facturas")
    return out


def proc_moviles(wb):
    """Hoja 'Servicios Móviles' → D.moviles[]
    Schema: {f, net, fo, pr, tipo, br, st}
    """
    rows = sheet_as_dicts(wb, "Servicios Móviles", "Servicios Moviles")
    out = []
    for r in rows:
        f   = to_ddmmyyyy(r.get("Fecha Emisión") or r.get("Fecha emision") or r.get("Fecha") or "")
        net = to_int(r.get("Monto Neto") or r.get("monto neto") or 0)
        if net == 0:
            continue
        out.append({
            "f":    f,
            "net":  net,
            "fo":   clean(r.get("N° Factura") or r.get("Factura") or ""),
            "pr":   clean(r.get("Proveedor") or ""),
            "tipo": clean(r.get("Tipo de Servicio") or r.get("Item ") or r.get("Item") or ""),
            "br":   to_int(r.get("Monto Mora (Bruto)") or r.get("Monto Bruto") or r.get("Monto bruto") or 0),
            "st":   clean(r.get("status") or r.get("Status") or ""),
        })
    print(f"    ✓ Servicios Móviles: {len(out)} facturas")
    return out


def proc_sms(wb):
    """Hoja 'Servicios SMS' → D.sms[]
    Schema: {fe, net, fo, pr, desc, cant, pu, ah, st}
    Nota: usa 'fe' (no 'f') como fecha — consistente con el portal.
    """
    rows = sheet_as_dicts(wb, "Servicios SMS")
    out = []
    for r in rows:
        fe  = to_ddmmyyyy(r.get("Fecha de emisión") or r.get("Fecha Emisión") or r.get("Fecha") or "")
        net = to_int(r.get("Monto Neto") or r.get("monto neto") or 0)
        if net == 0:
            continue
        out.append({
            "fe":   fe,
            "net":  net,
            "fo":   clean(r.get("Folio Factura") or r.get("Folio") or ""),
            "pr":   clean(r.get("Proveedor") or ""),
            "desc": clean(r.get("Descripción") or r.get("Descripcion") or ""),
            "cant": to_int(r.get("Cantidad de Mensajes") or r.get("Cantidad") or 0),
            "pu":   round(to_num(r.get("Precio según negociacion ") or r.get("Precio según negociacion") or 0), 6),
            "ah":   0,  # Sin columna de ahorro explícita en esta hoja
            # Linea base SMS = cantidad de mensajes (el ahorro se calcula con la
            # misma cantidad cambiando el precio) — acordado en reunion 04/08/2026
            "lb":   to_int(r.get("Linea Base") or r.get("Línea Base") or r.get("LB") or 0),
            "st":   clean(r.get("status") or r.get("Status") or ""),
        })
    print(f"    ✓ Servicios SMS: {len(out)} facturas")
    return out


def proc_fija(wb):
    """Hoja 'Servicios Telefonía Fija' → D.fija[]
    Schema: {f, m, net, fo, pr, desc, br, ah, st, fp}
    Nota: r.m se usa como 'YYYY-MM' para filtrar por mes.
    """
    rows = sheet_as_dicts(wb, "Servicios Telefonía Fija", "Servicios Telefonia Fija", "TELEFONÍA FIJA", "Servicios Telefonía IP", "OPS")
    out = []
    for r in rows:
        f_val = r.get("Fecha Emisión") or r.get("Fecha emision") or r.get("Fecha") or ""
        f   = to_ddmmyyyy(f_val)
        mes_val = r.get("Mes") or f_val
        m   = to_yyyymm(mes_val)
        net = to_int(r.get("Monto Neto ") or r.get("Monto Neto") or r.get("NETO") or r.get("Valor Neto") or 0)
        if net == 0:
            continue
        out.append({
            "f":    f,
            "m":    m,
            "net":  net,
            "fo":   clean(r.get("Folio Factura") or r.get("Folio") or r.get("Numero Factura") or r.get("Nro Factura") or r.get("N° facturación") or ""),
            "pr":   clean(r.get("Proveedor") or ""),
            "desc": clean(r.get("Descripción ") or r.get("Descripción") or r.get("Descripcion") or r.get("Detalle Facturación") or r.get("Tipo de servicio") or ""),
            "br":   to_int(r.get("Monto bruto") or r.get("Monto Bruto") or r.get("Valor facturado") or 0),
            "ah":   to_int(r.get("AHORRO MENSUAL") or r.get("Ahorro MES") or r.get("Ahorros Mensuales") or 0),
            "st":   clean(r.get("status") or r.get("Status") or r.get("ESTADO") or r.get("APROBADA / RECHAZADA") or ""),
            "fp":   clean(r.get("FECHA PAGO") or ""),
        })
    print(f"    ✓ Telefonía Fija: {len(out)} facturas")
    return out


def proc_fija600(wb):
    """Hoja 'Servicio Telefonia 600_809' → D.fija600[]
    Schema: {cuenta, f, fo, pr, desc, net, br, ah, st, fp}
    """
    rows = sheet_as_dicts(wb, "Servicio Telefonia 600_809", "Servicio sip trunk")
    out = []
    for r in rows:
        f   = to_ddmmyyyy(r.get("Fecha Emisión") or r.get("Fecha emision") or r.get("Fecha") or "")
        net = to_int(r.get("Monto Neto ") or r.get("Monto Neto") or 0)
        if net == 0:
            continue
        out.append({
            "cuenta": clean(r.get("Cuentas Facturación") or r.get("Cuentas Facturacion") or r.get("Nro. De Cliente") or ""),
            "f":      f,
            "fo":     clean(r.get("Folio Factura") or r.get("Folio") or r.get("Nro Factura") or ""),
            "pr":     clean(r.get("Proveedor") or ""),
            "desc":   clean(r.get("Descripción ") or r.get("Descripción") or r.get("Descripcion") or ""),
            "net":    net,
            "br":     to_int(r.get("Monto bruto") or r.get("Monto Bruto") or 0),
            "ah":     to_int(r.get("AHORRO MENSUAL") or 0),
            "st":     clean(r.get("status") or r.get("Status") or r.get("ESTADO") or ""),
            "fp":     clean(r.get("FECHA PAGO") or ""),
        })
    print(f"    ✓ Telefonía 600/809: {len(out)} facturas")
    return out


def proc_totalpack(wb):
    """Hoja 'TotalPack' → D.totalpack[]
    Schema: {fo, f, m, net, br, eq, ah, st}
    Los nombres de columna son variables (19 columnas), se detectan por patrón.
    """
    rows = sheet_as_dicts(wb, "TotalPack", "Servicios CCTV")
    out = []
    for r in rows:
        f_val = net = br = ah = 0
        fo = m_val = eq = st = ""
        f_val = None

        for k, v in r.items():
            kl = k.lower().strip()
            if f_val is None and ("fecha" in kl or "emisión" in kl or "emision" in kl):
                f_val = v
            if "mes" == kl or "mes consumo" in kl or "periodo" in kl:
                m_val = v
            if net == 0 and "neto" in kl and ("monto" in kl or "valor" in kl):
                net = to_int(v)
            if br == 0 and "bruto" in kl:
                br = to_int(v)
            if ah == 0 and "ahorro" in kl:
                ah = to_int(v)
            if not fo and ("folio" in kl or ("factura" in kl and "n°" not in kl)):
                fo = clean(v)
            if not st and "status" in kl:
                st = clean(v)
            if not eq and ("item" in kl or "equipo" in kl or "descripci" in kl):
                eq = clean(v)

        if net == 0:
            continue

        f = to_ddmmyyyy(f_val) if f_val else ""
        m = to_yyyymm(m_val) if m_val else to_yyyymm(f_val) if f_val else ""

        out.append({"fo": fo, "f": f, "m": m, "net": net, "br": br, "eq": eq, "ah": ah, "st": st})

    print(f"    ✓ TotalPack: {len(out)} registros")
    return out


def proc_impresoras(wb):
    """Hoja 'ServicioImpresoras' → D.impresoras[]
    Schema: {fo, f, m, net, br, eq, pr, ah, st}
    Ahorro calculado como LB - Monto neto si LB existe.
    """
    rows = sheet_as_dicts(wb, "ServicioImpresoras", "Servicios Impresión")
    out = []
    for r in rows:
        f_val = r.get("Fecha Emisión") or r.get("Fecha emision") or r.get("Fecha") or ""
        f   = to_ddmmyyyy(f_val)
        m_val = r.get("mes consumo") or r.get("Mes consumo") or r.get("Mes") or f_val
        m   = to_yyyymm(m_val)
        net = to_int(r.get("Monto neto") or r.get("Monto Neto") or 0)
        if net == 0:
            continue
        lb  = to_int(r.get("LB") or 0)
        br  = to_int(r.get("Monto Bruto") or r.get("Monto bruto") or 0)
        ah  = to_int(r.get("AHORRO") or r.get("Ahorros Mensuales") or 0) or (max(0, lb - net) if lb > 0 else 0)
        out.append({
            "fo":  clean(r.get("N° Factura") or r.get("Factura") or ""),
            "f":   f,
            "m":   m,
            "net": net,
            "br":  br,
            "eq":  clean(r.get("Item ") or r.get("Item") or r.get("Tipo de servicio") or ""),
            "pr":  clean(r.get("proveedor") or r.get("Proveedor") or ""),
            "ah":  ah,
            "st":  clean(r.get("status") or r.get("Status") or ""),
        })
    print(f"    ✓ Impresoras: {len(out)} facturas")
    return out


# ══════════════════════════════════════════════════════════════════════════════
# AUTENTICACIÓN Y DESCARGA SHAREPOINT
# ══════════════════════════════════════════════════════════════════════════════

def get_token():
    try:
        import requests
    except ImportError:
        print("  ⚠  Instala: pip install requests")
        return None

    if not CLIENT_SECRET:
        print("  ✗ AZURE_CLIENT_SECRET no está configurado.")
        print("    → Crea un .env con esa variable o expórtala antes de correr el script.")
        return None

    resp = requests.post(
        f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token",
        data={
            "client_id":     CLIENT_ID,
            "client_secret": CLIENT_SECRET,
            "scope":         "https://graph.microsoft.com/.default",
            "grant_type":    "client_credentials",
        },
        timeout=20,
    )
    if resp.status_code != 200:
        print(f"  ✗ Error de autenticación [{resp.status_code}]: {resp.text[:300]}")
        return None
    print("  ✓ Token Microsoft Graph obtenido")
    return resp.json()["access_token"]


def download_excel():
    try:
        import requests
        from urllib.parse import quote
    except ImportError:
        return None

    print("  🔑 Autenticando con Microsoft Graph...")
    token = get_token()
    if not token:
        return None

    headers = {"Authorization": f"Bearer {token}"}

    # Paso 1: Resolver site ID
    print("  🌐 Obteniendo site ID de SharePoint...")
    r_site = requests.get(
        f"https://graph.microsoft.com/v1.0/sites/{SP_SITE}",
        headers=headers, timeout=30
    )
    if r_site.status_code != 200:
        print(f"  ✗ Error obteniendo site: {r_site.status_code} — {r_site.text[:200]}")
        return None
    site_id = r_site.json()["id"]
    print(f"  ✓ Site ID obtenido")

    # Paso 2: Buscar la biblioteca por nombre
    print(f"  📚 Buscando biblioteca '{SP_LIBRARY}'...")
    r_drives = requests.get(
        f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives",
        headers=headers, timeout=30
    )
    drive_id = None
    if r_drives.status_code == 200:
        drives = r_drives.json().get("value", [])
        names  = [d.get("name", "") for d in drives]
        match  = next((d for d in drives if d.get("name") == SP_LIBRARY), None)
        if match:
            drive_id = match["id"]
            print(f"  ✓ Biblioteca encontrada: '{SP_LIBRARY}'")
        else:
            print(f"  ⚠ '{SP_LIBRARY}' no encontrada. Disponibles: {names}")
    else:
        print(f"  ⚠ No se pudieron listar drives: {r_drives.status_code}")

    # Paso 3: Descargar archivo
    encoded = quote(SP_FILE_PATH, safe="/")
    if drive_id:
        url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{encoded}:/content"
    else:
        # Fallback: drive por defecto con biblioteca/archivo combinados
        encoded_full = quote(f"{SP_LIBRARY}/{SP_FILE_PATH}", safe="/")
        url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{encoded_full}:/content"

    print("  📥 Descargando Excel desde SharePoint...")
    resp = requests.get(url, headers=headers, timeout=120, allow_redirects=True)

    if resp.status_code == 200:
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        tmp.write(resp.content)
        tmp.close()
        print(f"  ✓ Excel descargado ({len(resp.content) // 1024} KB)")
        return Path(tmp.name)
    else:
        print(f"  ✗ Error [{resp.status_code}]: {resp.text[:300]}")
        return None


# ══════════════════════════════════════════════════════════════════════════════
# GENERAR data.js
# ══════════════════════════════════════════════════════════════════════════════

def write_data_js(portal_data, output_path):
    ts = datetime.now().strftime("%d/%m/%Y %H:%M")
    portal_data["_updatedAt"] = ts

    js = (
        f"// Auto-generado por sync_data.py — {ts}\n"
        f"// Fuente: SharePoint EFIKA → {Path(SP_FILE_PATH).name}\n"
        f"// NO editar manualmente — se sobreescribe en cada sync\n"
        f"const PORTAL_DATA={json.dumps(portal_data, ensure_ascii=False, separators=(',', ':'))};\n"
    )

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(js)

    size_kb = output_path.stat().st_size // 1024
    total_rows = sum(len(v) for v in portal_data.values() if isinstance(v, list))
    print(f"  ✓ data.js generado — {total_rows:,} registros — {size_kb} KB")


# ══════════════════════════════════════════════════════════════════════════════
# GIT PUSH
# ══════════════════════════════════════════════════════════════════════════════

def git_push(folder):
    ts = datetime.now().strftime("%d/%m/%Y %H:%M")
    for cmd in [
        ["git", "-C", str(folder), "add", "data.js"],
        ["git", "-C", str(folder), "commit", "-m", f"Auto-sync SharePoint {ts}"],
        ["git", "-C", str(folder), "push"],
    ]:
        r = subprocess.run(cmd, capture_output=True, text=True)
        label = " ".join(cmd[2:4])
        out = (r.stderr or r.stdout or "").strip()
        if r.returncode != 0:
            if "nothing to commit" in out:
                print(f"  ℹ  git {label}: sin cambios")
            else:
                print(f"  ⚠  git {label}: {out[:120]}")
        else:
            print(f"  ✓  git {label}")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    args = sys.argv[1:]
    do_push   = "--push" in args
    local_arg = next((a for a in args if not a.startswith("--")), None)

    print("\n╔══════════════════════════════════════════════════════════╗")
    print("║  EFIKA Portal — Sync SharePoint → data.js                ║")
    print("╚══════════════════════════════════════════════════════════╝\n")

    folder = Path(__file__).parent
    output = folder / "data.js"

    # ── 1. Obtener Excel ──────────────────────────────────────────────────────
    tmp_xl = None
    if local_arg and Path(local_arg).exists():
        xl_path = Path(local_arg)
        print(f"  📂 Modo local: {xl_path.name}")
    else:
        xl_path = None
        # Buscar Excel local en la carpeta
        for pat in ["Control Consolidado*.xlsx", "*.xlsx"]:
            found = [f for f in folder.glob(pat) if not f.name.startswith("~")]
            if found:
                xl_path = found[0]
                print(f"  📂 Excel local: {xl_path.name}")
                break
        if xl_path is None:
            xl_path = download_excel()
            tmp_xl = xl_path

    if xl_path is None:
        print("  ✗ No se pudo obtener el Excel.")
        print("    → Verifica AZURE_CLIENT_SECRET y la ruta del archivo en SharePoint.")
        sys.exit(1)

    # ── 2. Leer Excel ─────────────────────────────────────────────────────────
    try:
        import openpyxl
    except ImportError:
        print("  ⚠  Instala: pip install openpyxl")
        sys.exit(1)

    print(f"\n  📊 Procesando hojas de {xl_path.name}...")
    try:
        wb = openpyxl.load_workbook(str(xl_path), data_only=True, read_only=True)
    except Exception as e:
        print(f"  ✗ Error abriendo Excel: {e}")
        sys.exit(1)

    print(f"  📋 Hojas: {wb.sheetnames}\n")

    # ── DEBUG: estructura de cada hoja (para ajustar el parser por cliente) ───
    for _name in wb.sheetnames:
        try:
            _ws = wb[_name]
            _hdr = None
            for _row in _ws.iter_rows(values_only=True):
                if sum(1 for c in _row if c is not None and str(c).strip()) >= 3:
                    _hdr = [clean(c) for c in _row][:16]
                    break
            print(f"  🔎 '{_name}' → {_hdr}")
        except Exception as _e:
            print(f"  🔎 '{_name}' → error: {_e}")
    print()

    resumen    = proc_resumen(wb)
    privados   = proc_privados(wb)
    moviles    = proc_moviles(wb)
    sms        = proc_sms(wb)
    fija       = proc_fija(wb)
    fija600    = proc_fija600(wb)
    totalpack  = proc_totalpack(wb)
    impresoras = proc_impresoras(wb)

    # ── 3. Generar data.js ────────────────────────────────────────────────────
    portal_data = {
        "resumen":    resumen,
        "privados":   privados,
        "moviles":    moviles,
        "sms":        sms,
        "fija":       fija,
        "fija600":    fija600,
        "totalpack":  totalpack,
        "impresoras": impresoras,
    }

    # ── Fallback: si no hay hoja de Resumen, construirlo desde las facturas ──
    if not resumen:
        from collections import defaultdict
        agg = defaultdict(lambda: {"mv":0,"sms":0,"pv":0,"fi":0,"ms":0})
        def _mes(r):
            m = r.get("m") or ""
            if m and len(str(m)) >= 7: return str(m)[:7]
            f = r.get("f") or r.get("fe") or ""
            p = str(f).split("/")
            return f"{p[2]}-{p[1]}" if len(p) == 3 else ""
        for key, rows in (("pv", privados), ("mv", moviles), ("sms", sms), ("fi", fija)):
            for r in rows:
                k = _mes(r)
                if k: agg[k][key] += int(r.get("ah") or 0) or 0
        # si no hay ahorros informados, usar el neto facturado como referencia de gestión
        sin_ahorro = all(sum(v.values()) == 0 for v in agg.values()) if agg else True
        if sin_ahorro:
            agg = defaultdict(lambda: {"mv":0,"sms":0,"pv":0,"fi":0,"ms":0})
            for key, rows in (("pv", privados), ("mv", moviles), ("sms", sms), ("fi", fija)):
                for r in rows:
                    k = _mes(r)
                    if k: agg[k][key] += int(r.get("net") or 0)
        for k in sorted(agg):
            v = agg[k]; tot = sum(v.values())
            resumen.append({"m": k + "-01", "total": tot, "mv": v["mv"], "sms": v["sms"],
                            "pv": v["pv"], "fi": v["fi"], "ms": v["ms"], "ah": tot})
        if resumen:
            print(f"    ✓ Resumen calculado desde facturas: {len(resumen)} meses")
        portal_data["resumen"] = resumen

    print(f"\n  📝 Generando data.js...")
    write_data_js(portal_data, output)

    # Limpiar temporal
    if tmp_xl and Path(tmp_xl).exists():
        Path(tmp_xl).unlink()

    # ── 4. Git push ───────────────────────────────────────────────────────────
    if do_push:
        print(f"\n  📤 Subiendo a GitHub Pages...")
        git_push(folder)
        print(f"  🌐 Portal se actualiza en ~60 segundos")
    else:
        print(f"\n  💡 Para subir a GitHub Pages: python sync_data.py --push")

    # ── Resumen ───────────────────────────────────────────────────────────────
    if resumen:
        last = resumen[-1]
        total_ah = sum(r["total"] for r in resumen)
        print(f"\n  📊 Último mes:        {last['m'][:7]}")
        print(f"  💰 Ahorro acumulado:  ${total_ah:,.0f}")
        print(f"  📄 Facturas privados: {len(privados)}")
        print(f"  📱 Facturas móviles:  {len(moviles)}")
        print(f"  💬 Facturas SMS:      {len(sms)}")
        print(f"  ☎️  Facturas fija:     {len(fija)} + {len(fija600)} (600/809)")

    print("\n  ✅ Completado\n")


if __name__ == "__main__":
    main()
